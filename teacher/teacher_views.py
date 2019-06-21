import os
import random
import time
from datetime import datetime
from django.utils.http import urlquote
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.db import connection
import openpyxl
import json
# Create your views here.
from python_end.Student import Student
from python_end.Course import Course
from python_end.Participation import Participation
from django.utils.encoding import escape_uri_path
import os
def get_filename(request):
    adict = get_students(request)
    names=[]
    for i, j, k in os.walk(request.session.get('teacherid')):
        adict['names']=k
        adict['warecount']=len(k)
    print(adict)
    return adict

def get_participation_status(request):
    cursor = connection.cursor()
    sql="select * from participation where start_date is not null and end_date is null and course_id in (select id from course where teacher_id=%s)"
    cursor.execute(sql,request.session.get('teacherid'))
    a=cursor.fetchall()
    if len(a)==0:
        return -1
    return 1
#打包返回教师信息的方法
def get_teacher_dict(request):
    teacherid = request.session.get('teacherid')
    teachername = request.session.get('teachername')
    course = getAllCourse(request)
    return {'teacherid':teacherid,'teachername':teachername,"courses":course}

def get_now_class(request):
    cursor = connection.cursor()
    teacherid = request.session.get('teacherid')
    sql="select * from `now-course`,`Course` where start_date is not null and end_date is null and `now-course`.course_id =Course.id and teacher_id=%s"
    cursor.execute(sql,teacherid)
    a=cursor.fetchall()
    if len(a)==0:
        return None
    else:
        return Course(a[0][1],a[0][6]),a[0][2]
def get_students(request):
    adict = get_teacher_dict(request)
    adict['nowclick'] = int(request.session.get('cid'))
    cursor = connection.cursor()
    sql = "SELECT count(*) FROM `student-course` where course_id=%s"
    cursor.execute(sql, request.session.get('cid'))
    c = cursor.fetchall()
    adict['studentcount'] = int(c[0][0])
    students = []
    sql = "SELECT * FROM `student-course` where course_id=%s"
    cursor.execute(sql, request.session.get('cid'))
    c = cursor.fetchall()
    adict['list'] = [int(x) for x in range(1, len(c))]
    for i in range(len(c)):
        sql = "select count(*),sum(score) from `student-course`,`course-student` where `student-course`.id=`course-student`.course_selectID and " \
              "`student-course`.course_id=%s and  `student-course`.student_id=%s"
        cursor.execute(sql, (request.session.get('cid'), c[i][1]))
        e = cursor.fetchall()

        sql = "select count(*) from `participation`,`student-participation` where `participation`.id=`student-participation`.participation_id " \
              "and participation.course_id=%s and student_id=%s"
        cursor.execute(sql, (request.session.get('cid'),c[i][1]))
        d = cursor.fetchall()

        s = Student(c[i][1], c[i][2], d[0][0], e[0][0], e[0][1])
        students.append(s)
    students = sorted(students, key=lambda x: int(x.ID))
    adict['student'] = students
    sql="select count(*) from Participation where end_date is not null and course_id=%s"
    cursor.execute(sql, request.session.get('cid'))
    sb=cursor.fetchall()
    adict['partcount']=int(sb[0][0])
    return adict
######################################

def index(request):
    teacherid=request.session.get('teacherid')
    teachername=request.session.get('teachername')
    course=getAllCourse(request)
    adict=get_teacher_dict(request)
    adict['participationstatus']=get_participation_status(request)
    try:
        a,b=get_now_class(request)
    except:
        pass
    if get_now_class(request) is not None:
        request.session['cid']=a.ID
        adict['nowcourse'],start_time=get_now_class(request)
        adict['timestamp']=int(time.time())-int(time.mktime(start_time.timetuple()))
    else:
        adict['nowcourse'] =Course(-1,'无')
    if teacherid:
        return render(request,'teacher/teacherhome.html',adict)
    else:
        return redirect('/index.html')
def getAllCourse(request):
    cursor = connection.cursor()
    tablename='Course'
    teacher_id=request.session['teacherid']
    print(teacher_id)
    sql = "select * from Course where teacher_id= %s"
    cursor.execute(sql,teacher_id)
    c = cursor.fetchall()
    courses=[]
    for i in c:
        courses.append(Course(i[0],i[2]))
    return courses
def addcourseaction(request):
    sql="insert into Course (teacher_id,name) values(%s,%s)"
    cursor = connection.cursor()
    cursor.execute(sql,(request.session.get('teacherid'),request.POST['classname']))

    courses=getAllCourse(request)
    file=request.FILES.get("file", None)
    wb = openpyxl.load_workbook(file)
    sheet=wb.active
    for i in range(1,sheet.max_row+1):
        sql2 = "insert into `student-course` (student_id,student_name,course_id) values(%s,%s,%s)"
        cursor.execute(sql2, (sheet.cell(i,1).value,sheet.cell(i,2).value,courses[len(courses) - 1].ID))
    return redirect('/teacher/coursedetail?id='+str(courses[len(courses) - 1].ID))

def coursedetailview(request):
    adict=get_teacher_dict(request)
    adict['nowclick']=int(request.GET['id'])
    cursor = connection.cursor()
    sql = "SELECT count(*) FROM `student-course` where course_id=%s"
    cursor.execute(sql,request.GET['id'])
    c = cursor.fetchall()
    adict['studentcount']=int(c[0][0])
    sql="select count(*) from Participation where end_date is not null and course_id=%s"
    cursor.execute(sql, request.GET['id'])
    d=cursor.fetchall()
    adict['partcount']=int(d[0][0])
    request.session['cid']=request.GET['id']
    print (adict)
    return render(request,'teacher/courseinfo.html',adict)
def addcourseview(request):
    return render(request, 'teacher/addcourse.html',get_teacher_dict(request))

def deletecourse(request):
    cursor = connection.cursor()

    sql="delete FROM `course-student` where course_selectID in(select id from `student-course` where course_id=%s)"
    cursor.execute(sql, (request.POST['opid']))

    sql1 = "delete  FROM `student-course` where course_id=%s;"
    cursor.execute(sql1,(request.POST['opid']))

    sql2 = "delete  FROM `Student-Participation` where participation_id in (select id from `Participation` where course_id=%s) ;"
    cursor.execute(sql2, (request.POST['opid']))

    sql3 = "delete  FROM `Participation` where course_id=%s;"
    cursor.execute(sql3, (request.POST['opid']))

    sql4="delete  FROM `course` where id=%s;"
    cursor.execute(sql4, (request.POST['opid']))

    sql5 = "delete FROM `now-course` where course_id=%s"
    cursor.execute(sql5, (request.POST['opid']))
    return redirect('index.html')
def studentdetailview(request):
    return render(request, 'teacher/studentdetail.html', get_students(request))
def downstu(request):
    filename = time.strftime("%Y-%m-%d-")
    id=request.GET['id']
    teacher_id=request.session['teacherid']
    cursor = connection.cursor()
    sql = "select name from Course where teacher_id= %s and id=%s"
    cursor.execute(sql,(teacher_id,id))
    c = cursor.fetchall()
    filename+=c[0][0]+"成绩汇总.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(1,1).value='学号'
    sheet.cell(1, 2).value = '姓名'
    sheet.cell(1, 3).value = '出勤次数'
    sheet.cell(1, 4).value = '老师提问次数'
    sheet.cell(1, 5).value = '提问总分'
    a=get_students(request)
    students=a["student"]
    for i in range(len(students)):
        sheet.cell(i+2, 1).value = students[i].ID
        sheet.cell(i+2, 2).value = students[i].name
        sheet.cell(i+2, 3).value = students[i].pcount
        sheet.cell(i+2, 4).value = students[i].qcount
        sheet.cell(i+2, 5).value = students[i].qscore
    wb.save("temp/"+filename)
    with open('temp/'+filename, 'rb') as model_excel:
        result = model_excel.read()
    response = HttpResponse(result)
    response['Content-Disposition'] = "attachment; filename*=utf-8''{}".format(escape_uri_path(filename))
    os.remove("temp/"+filename)
    return response

def pardetailview(request):
    id = request.GET['id']
    cursor = connection.cursor()
    sql='select * from Participation where end_date is not null and course_id=%s'
    cursor.execute(sql,id)
    c=cursor.fetchall()

    s='select count(*) from `Student-Course` where course_id=%s'
    cursor.execute(s,id)
    all=cursor.fetchall()
    ps=[]
    for i in range(len(c)):
        sql2 = 'select count(*) from `Student-Participation` where participation_id=%s'
        cursor.execute(sql2, c[i][0])
        d = cursor.fetchall()
        ps.append(Participation(c[i][2],c[i][3],d[0][0],int(d[0][0])/int(all[0][0])))
    adict=get_students(request)
    adict['par']=ps
    return render(request, 'teacher/participation.html', adict)

def start_class(request):
    if not get_now_class(request):
        cursor = connection.cursor()
        id=request.GET['id']
        sql='insert into `now-course` (course_id,start_date)values(%s,now())'
        cursor.execute(sql, id)
        return redirect('/teacher/index.html')
    else:
        return redirect('/teacher/index.html')
def end_class(request):
    if get_now_class(request):
        cursor = connection.cursor()
        id = request.GET['id']
        sql = 'update `now-course` set end_date=now() where course_id=%s and end_date is null'
        cursor.execute(sql, id)
        return redirect('/teacher/index.html')
    else:
        return redirect('/teacher/index.html')
def start_participation(request):
    cursor = connection.cursor()
    sql="insert into participation (course_id,start_date) values(%s,now())"
    cursor.execute(sql,request.GET['id'])
    return redirect('/teacher/index.html')
def end_participation(request):
    cursor = connection.cursor()
    sql = "update participation set end_date=now() where course_id=%s and end_date is null and start_date is not null"
    cursor.execute(sql, request.GET['id'])
    return redirect('/teacher/index.html')
def coursewareview(request):
    return render(request, 'teacher/courseware.html', get_filename(request))
def upload(request):
    file = request.FILES.get('file')
    if not os.path.exists(request.session.get('teacherid')+"/"):
        os.makedirs(request.session.get('teacherid')+"/")
    with open(request.session.get('teacherid')+"/"+file.name,'wb') as f:
        for i in file.readlines():
            f.write(i)
    return redirect('/teacher/courseware?id='+request.session.get('cid'))
def deleteware(request):
    os.remove(request.session.get('teacherid') + "/"+request.GET['name'])
    return redirect('/teacher/courseware?id=' + request.session.get('cid'))
def downware(request):
    file = open(request.session.get('teacherid')+"/"+request.GET['name'], 'rb')
    response = HttpResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = "attachment;filename*=utf-8''{}".format(escape_uri_path(request.GET['name']))
    return response
def classware(request):
    teacherid = request.session.get('teacherid')
    teachername = request.session.get('teachername')
    course = getAllCourse(request)
    a=get_filename(request)
    adict = get_teacher_dict(request)
    adict.update(a)
    adict['participationstatus'] = get_participation_status(request)
    adict['index']='active'
    if get_now_class(request) is not None:
        adict['nowcourse'], start_time = get_now_class(request)
        adict['timestamp'] = int(time.time()) - int(time.mktime(start_time.timetuple()))
    else:
        adict['nowcourse'] = Course(-1, '无')
    if teacherid:
        return render(request, 'teacher/downware.html', adict)
    else:
        return redirect('/index.html')
def randpick(request):
    cursor = connection.cursor()
    sql = "SELECT * FROM `student-course` where course_id=%s"
    cursor.execute(sql, request.GET['id'])
    a=cursor.fetchall()
    alist=[]
    for i in a:
        s={'sid':i[0],'id':i[1],'name':i[2],'cid':i[3]}
        alist.append(s)
    person=random.sample(alist, 1)
    print(person)
    return render(request,'teacher/pick.html',{'student':person[0]})
def setscore(request):
    print("成绩"+request.GET['score'])
    print("选课编号"+request.GET['sid'])
    cursor = connection.cursor()
    sql = "insert into  django_stu_system.`course-student` values(%s,now(),%s)";
    cursor.execute(sql, (request.GET['sid'],request.GET['score']))
    return redirect('/teacher/index.html')
def getquestion(request):
    cid = request.GET['id']
    cursor = connection.cursor()
    sql = "select * from question where course_id=%s"
    cursor.execute(sql, cid)
    a = cursor.fetchall()
    alist = []
    for i in a:
        sql2 = "select name from student where id=%s"
        cursor.execute(sql2, i[1])
        b = cursor.fetchall()
        adict = {'qid': i[0], 'name': b[0][0], 'content': i[3]}
        alist.append(adict)

    teacherid = request.session.get('teacherid')
    teachername = request.session.get('teachername')
    course = getAllCourse(request)
    a = get_filename(request)
    adict = get_teacher_dict(request)
    adict['questions']=alist
    adict.update(a)
    adict['participationstatus'] = get_participation_status(request)
    adict['index'] = 'active'
    if get_now_class(request) is not None:
        adict['nowcourse'], start_time = get_now_class(request)
        adict['timestamp'] = int(time.time()) - int(time.mktime(start_time.timetuple()))
    else:
        adict['nowcourse'] = Course(-1, '无')
    if teacherid:
        return render(request, 'teacher/getquestion.html', adict)
    else:
        return redirect('/index.html')
def deletequestion(request):
    qid=request.GET['qid']
    cursor = connection.cursor()
    sql = "delete  from question where id=%s"
    cursor.execute(sql, qid)
    return redirect('/teacher/getquestion?id='+str(request.session.get('cid')))




